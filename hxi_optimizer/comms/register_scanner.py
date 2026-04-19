"""Auto-scan register map from Steve's Register_List.xlsx.

Parses all 10 sections, resolves WORD bit-fields, REAL (FLOAT32) spans,
and INT registers into a unified register catalogue that the optimizer,
dashboard, and training pipeline all use as the single source of truth.

Usage:
    from hxi_optimizer.comms.register_scanner import load_register_catalog
    catalog = load_register_catalog("Register_List.xlsx")
"""
from __future__ import annotations

import logging
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

logger = logging.getLogger("register_scanner")


@dataclass
class Register:
    address: int            # pymodbus 0-indexed (GE %RNNNN - 1)
    ge_address: str         # original "%R06600" string
    name: str               # human-readable ("RPM Real Encoder")
    original: str           # PLC-internal reference ("%R02760")
    section: str            # sheet name ("Smart Slide")
    dtype: str              # "REAL" | "INT" | "WORD"
    bit: Optional[int] = None   # for WORD bit-fields: bit position 0..15
    word_count: int = 1     # 2 for REAL (FLOAT32), 1 for INT/WORD
    is_spare: bool = False
    writable: bool = False  # True for registers the optimizer may write


@dataclass
class RegisterCatalog:
    """Complete register map parsed from Excel."""
    registers: list[Register] = field(default_factory=list)
    by_name: dict[str, Register] = field(default_factory=dict)
    by_address: dict[int, Register] = field(default_factory=dict)
    sections: dict[str, list[Register]] = field(default_factory=dict)

    # Computed read blocks for efficient Modbus FC03 transactions
    read_blocks: list[dict] = field(default_factory=list)

    def get(self, name: str) -> Optional[Register]:
        return self.by_name.get(name)

    def address_of(self, name: str) -> Optional[int]:
        r = self.by_name.get(name)
        return r.address if r else None

    def non_spare(self) -> list[Register]:
        return [r for r in self.registers if not r.is_spare]

    def in_section(self, section: str) -> list[Register]:
        return self.sections.get(section, [])

    def word_registers(self) -> list[Register]:
        """All WORD registers with bit fields (for digital I/O decoding)."""
        return [r for r in self.registers if r.dtype == "WORD" and r.bit is not None]

    def real_registers(self) -> list[Register]:
        return [r for r in self.registers if r.dtype == "REAL"]

    def int_registers(self) -> list[Register]:
        return [r for r in self.registers if r.dtype == "INT"]

    def compute_read_blocks(self, max_gap: int = 10) -> list[dict]:
        """Compute optimal FC03 read blocks (merge nearby registers)."""
        addrs = sorted(set(r.address for r in self.registers if not r.is_spare))
        if not addrs:
            return []
        blocks = []
        block_start = addrs[0]
        block_end = addrs[0]
        for addr in addrs[1:]:
            wc = next((r.word_count for r in self.registers if r.address == addr), 1)
            if addr - block_end <= max_gap:
                block_end = addr + wc - 1
            else:
                blocks.append({"start": block_start,
                               "count": block_end - block_start + 1})
                block_start = addr
                block_end = addr + wc - 1
        blocks.append({"start": block_start,
                       "count": block_end - block_start + 1})
        self.read_blocks = blocks
        return blocks


_ADDR_RE = re.compile(r"%R(\d+)(?:\s*\((\d+)\))?")

WRITABLE_NAMES = {
    "Swash Lower Threshold", "Swash Upper Limit",
    "FWD DEGREES BUMP SET", "REV DEGREES BUMP SET",
}


def _parse_address(addr_str: str) -> tuple[int, Optional[int]]:
    """Parse '%R06600' or '%R06664 (3)' → (pymodbus_addr, bit_or_None)."""
    m = _ADDR_RE.match(addr_str.strip())
    if not m:
        raise ValueError(f"Cannot parse address: {addr_str!r}")
    ge_num = int(m.group(1))
    bit = int(m.group(2)) if m.group(2) else None
    return ge_num - 1, bit  # pymodbus 0-indexed


def load_register_catalog(xlsx_path: str | Path) -> RegisterCatalog:
    """Parse Register_List.xlsx into a RegisterCatalog."""
    import openpyxl
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    catalog = RegisterCatalog()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        section_regs = []
        for row in ws.iter_rows(min_row=1, values_only=True):
            if not row or len(row) < 4:
                continue
            original, name, dtype, addr_str = row[:4]
            if not addr_str or not isinstance(addr_str, str) or not addr_str.startswith("%R"):
                continue
            if not dtype or str(dtype).strip().upper() not in ("REAL", "INT", "WORD", "INT "):
                continue

            name = str(name).strip() if name else "UNNAMED"
            original = str(original).strip() if original else ""
            dtype = str(dtype).strip().upper().rstrip()
            is_spare = "SPARE" in name.upper()

            try:
                address, bit = _parse_address(addr_str)
            except ValueError:
                logger.warning(f"Skipping unparseable address: {addr_str}")
                continue

            wc = 2 if dtype == "REAL" else 1
            writable = name in WRITABLE_NAMES

            reg = Register(
                address=address,
                ge_address=addr_str.strip(),
                name=name,
                original=original,
                section=sheet_name,
                dtype=dtype,
                bit=bit,
                word_count=wc,
                is_spare=is_spare,
                writable=writable,
            )
            catalog.registers.append(reg)
            section_regs.append(reg)

            # Index by a cleaned name key
            key = _make_key(name, sheet_name, bit)
            catalog.by_name[key] = reg
            if bit is None:
                catalog.by_address[address] = reg

        catalog.sections[sheet_name] = section_regs

    catalog.compute_read_blocks()
    logger.info(f"Loaded {len(catalog.registers)} registers from {xlsx_path} "
                f"({len(catalog.non_spare())} active, "
                f"{len(catalog.read_blocks)} read blocks)")
    return catalog


def _make_key(name: str, section: str, bit: Optional[int]) -> str:
    """Generate a unique dict key from register name."""
    key = re.sub(r"[^a-zA-Z0-9]+", "_", name).strip("_").lower()
    if bit is not None:
        key += f"_bit{bit}"
    if not key or key == "spare":
        key = f"{section.lower().replace(' ', '_')}_{key}"
    return key


def catalog_to_json(catalog: RegisterCatalog) -> list[dict]:
    """Serialize catalog for dashboard API."""
    out = []
    for r in catalog.registers:
        out.append({
            "address": r.address,
            "ge_address": r.ge_address,
            "name": r.name,
            "original": r.original,
            "section": r.section,
            "dtype": r.dtype,
            "bit": r.bit,
            "word_count": r.word_count,
            "is_spare": r.is_spare,
            "writable": r.writable,
        })
    return out


def catalog_to_read_plan(catalog: RegisterCatalog) -> list[dict]:
    """Generate Modbus read plan for the dashboard scanner."""
    return catalog.read_blocks
